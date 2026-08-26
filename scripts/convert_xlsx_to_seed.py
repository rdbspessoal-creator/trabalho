#!/usr/bin/env python3
"""
Converte as planilhas-fonte para os JSONs de seed usados pelo sistema:
  data/seed-clientes.json   <- CLIENTES_X_EQUIPE.xlsx (aba "Cliente X Equipe")
  data/seed-demandas.json   <- Demanda_de_Medição_Modelo.xlsx
                                (abas HISTÓRICO_ORIGINAL + MÊS_ATUAL)

Uso:
  python3 scripts/convert_xlsx_to_seed.py <Demanda_de_Medição_Modelo.xlsx> <CLIENTES_X_EQUIPE.xlsx>

Regras (replicando a especificação do sistema):
- HISTÓRICO_ORIGINAL / MÊS_ATUAL: cabeçalho real está na linha com "CLIENTE"
  na coluna B; linhas acima são anotações soltas e são ignoradas; linhas sem
  CLIENTE preenchido são separadores de data e também são ignoradas.
- EQUIPE -> ÁREA: "AUTOMAÇÃO NORTE"->NORTE, "AUTOMAÇÃO SUL"->SUL,
  "AUTOMAÇÃO OESTE" / "AUTOMAÇÃO OESTE/SERTÃO"->OESTE.
- EXECUTANTE -> ÁREA (para o histórico seed): contém "BRUNO"->NORTE,
  "RUBENING"->SUL, "ILTEARLE"->OESTE; qualquer outro (MEDIÇÃO, PEDRO,
  JOSINALDO, R. PEREIRA...) -> OUTRO (não correlacionado, é registro
  administrativo/histórico já fechado).
"""
import sys
import json
import re
import unicodedata
import openpyxl


def normalize(s):
    s = (s or "").upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def equipe_to_area(equipe):
    n = normalize(equipe)
    if "NORTE" in n:
        return "NORTE"
    if "SUL" in n:
        return "SUL"
    if "OESTE" in n or "SERTAO" in n:
        return "OESTE"
    return None


def exec_to_area(executante):
    n = normalize(executante)
    if "BRUNO" in n:
        return "NORTE"
    if "RUBENING" in n:
        return "SUL"
    if "ILTEARLE" in n:
        return "OESTE"
    return "OUTRO"


def parse_clientes(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Cliente X Equipe"]
    out = []
    for r in range(2, ws.max_row + 1):
        cliente = ws.cell(row=r, column=1).value
        equipe = ws.cell(row=r, column=2).value
        if not cliente:
            continue
        cliente = str(cliente).strip()
        equipe = (equipe or "").strip()
        area = equipe_to_area(equipe)
        codigo, _, nome = cliente.partition(" - ")
        out.append(
            {
                "cliente": cliente,
                "codigo": codigo.strip(),
                "nome": nome.strip() or cliente,
                "equipe": equipe,
                "area": area,
            }
        )
    return out


def parse_demanda_sheet(ws):
    # Acha a linha e a coluna de cabeçalho real (onde uma célula == "CLIENTE");
    # a posição varia entre abas (HISTÓRICO_ORIGINAL usa coluna B, MÊS_ATUAL usa coluna A).
    header_row = None
    header_col = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            if str(ws.cell(row=r, column=c).value or "").strip().upper() == "CLIENTE":
                header_row, header_col = r, c
                break
        if header_row is not None:
            break
    if header_row is None:
        return []
    cc = header_col
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        cliente = ws.cell(row=r, column=cc).value
        if not cliente or not str(cliente).strip():
            continue
        executante = ws.cell(row=r, column=cc + 1).value or ""
        data = ws.cell(row=r, column=cc + 2).value
        defeito = ws.cell(row=r, column=cc + 3).value or ""
        acao = ws.cell(row=r, column=cc + 4).value or ""
        detalhamento = ws.cell(row=r, column=cc + 5).value or ""
        data_iso = ""
        if hasattr(data, "strftime"):
            data_iso = data.strftime("%Y-%m-%d")
        elif data:
            data_iso = str(data).strip()
        out.append(
            {
                "cliente": str(cliente).strip(),
                "executante": str(executante).strip(),
                "area": exec_to_area(str(executante)),
                "data": data_iso,
                "defeito": str(defeito).strip(),
                "acao": str(acao).strip(),
                "detalhamento": str(detalhamento).strip(),
            }
        )
    return out


def parse_demandas(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    historico = parse_demanda_sheet(wb["HISTÓRICO_ORIGINAL"])
    mes_atual = parse_demanda_sheet(wb["MÊS_ATUAL"])
    return historico, mes_atual


# Lacunas de dados já resolvidas na especificação (seção 8): clientes do
# histórico que não batem bem com a base do Excel, mas cuja área é conhecida
# com evidência clara (aparecem sempre com o mesmo técnico/área no histórico).
KNOWN_GAP_FIXES = [
    {
        "cliente": "POSTO PRINCESA ISABEL",
        "area": "NORTE",
    },
    {
        "cliente": "MAURICEIA RAÇÕES",
        "area": "OESTE",
    },
    {
        # variantes de grafia citadas na especificação (seção 8)
        "cliente": "MAURICÉA RAÇÕES",
        "area": "OESTE",
    },
    {
        "cliente": "MAURICEA RAÇÕES",
        "area": "OESTE",
    },
]


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    demanda_path, clientes_path = sys.argv[1], sys.argv[2]

    clientes = parse_clientes(clientes_path)
    for fix in KNOWN_GAP_FIXES:
        clientes.append(
            {
                "cliente": fix["cliente"],
                "codigo": "",
                "nome": fix["cliente"],
                "equipe": f"AUTOMAÇÃO {fix['area']}",
                "area": fix["area"],
            }
        )

    historico, mes_atual = parse_demandas(demanda_path)

    with open("data/seed-clientes.json", "w", encoding="utf-8") as f:
        json.dump(clientes, f, ensure_ascii=False, indent=2)

    demandas_seed = {"historico": historico, "mesAtual": mes_atual}
    with open("data/seed-demandas.json", "w", encoding="utf-8") as f:
        json.dump(demandas_seed, f, ensure_ascii=False, indent=2)

    print(f"Clientes: {len(clientes)} (base {len(clientes) - len(KNOWN_GAP_FIXES)} + {len(KNOWN_GAP_FIXES)} lacunas resolvidas)")
    print(f"Histórico original: {len(historico)} registros")
    print(f"Mês atual: {len(mes_atual)} registros")
    print(f"Total seed de demandas: {len(historico) + len(mes_atual)} registros")


if __name__ == "__main__":
    main()
